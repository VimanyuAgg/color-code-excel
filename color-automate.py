# import xlsxwriter
# from xlsxwriter.workbook import Workbook 

# workbook = xlsxwriter.Workbook('consolidated_report.xlsx')
# print workbook
# print "workbook created"

# print workbook.worksheets()
# worksheet= workbook.get_worksheet_by_name("Consolidated_Report__crosstab")
# print worksheet
# print "worksheet created"
# green_format = workbook.add_format()
# green_format.set_pattern(1)
# green_format.set_bg_color('#008000')

# worksheet.write(1,5,green_format)


# worksheet = workbook.add_worksheet('Hello')
# workbook.close()

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import datetime, timedelta

wb = openpyxl.load_workbook(filename='consolidated_report.xlsx')
# ws = wb.get_worksheet_by_name('Consolidated_Report__crosstab')

ws = wb['Consolidated_Report__crosstab']

# print ws
# total_entry = ws['A']
# totalRows = 0
# for t in total_entry:
# 	if t.value == None:
# 		break
# 	else:
# 		totalRows +=1

# print totalRows
# print ws.max_row
#print ws['A1'].value
print "Starting up services...."
print ""

greenFill = PatternFill(start_color='10ff00',
                   end_color='10ff00',
                   fill_type='solid')

redFill = PatternFill(start_color='FFFF0000',
				   end_color='FFFF0000',
                   fill_type='solid')

amberFill = PatternFill(start_color='FFC200',
                   end_color='FFC200',
                   fill_type='solid')
greyFill =  PatternFill(start_color='696969',
                   end_color='696969',
                   fill_type='solid')

wineFill =  PatternFill(start_color='722f37',
                   end_color='722f37',
                   fill_type='solid')

print "Color codes defined"
project_status = ws['F']
project_stage = ws['E']
project_per_complete = ws['N']
project_start_date =ws['J']
project_end_date = ws['K']
last_updated  = ws['M']
rag_status = ws['G']
rag_reason = ws['H']
margin_variance = ws['O']
delivery_days_variance = ws['P']
reason_overrun = ws['Q']
delivery_country = ws['T']
project_type= ws['AB']
# to_expiry_date = ws['J']
project_category = ws['AD']
project_revenue = ws['AE']
revenue_source = ws['Z']
amount_invoiced = ws['AF']
# project_reclass_value = ws['AG']
invoiced_method = ws['AA']
project_cost_budget = ws['AI']
project_margin_budget = ws['AO']
delivery_days_budget = ws['BI']
delivery_days_forecast = ws['BJ']
days_budget = ws['BN']
days_forecast = ws['BO']
project_reclass = ws['AG']
booked_date = ws['I']

print "Setting necessary variables... Beginning file analysis"
for i in range (1,len(project_status)):
	if (i == len(project_status)/2):
		print "Halfway done"
	# print project_per_complete[i].value > 1
	# print type(project_per_complete[i].value)


	# R9C1
	if project_start_date[i].value == None and project_status[i].value != "Booked":
		project_start_date[i].fill = redFill
		project_status[i].fill = greyFill


	# R10C1
	if project_end_date[i].value == None and project_status[i].value != "Booked":
		project_end_date[i].fill = redFill
		project_status[i].fill = greyFill

	# R10C4
	if (project_end_date[i].value != None and project_start_date[i].value != None) and project_end_date[i].value < project_start_date[i].value:
		project_end_date[i].fill = redFill
		project_start_date[i].fill = greyFill

	# R10C2
	# print type(project_end_date[i].value)
	if (project_end_date[i].value != None) and (project_end_date[i].value < datetime.now() and (project_status[i].value != "Complete" or (project_status[i].value.find("Complete") != -1 and project_status[i].value.find("-") != -1) or project_status[i].value != "Closed")):
		project_end_date[i].fill = redFill
		project_status[i].fill = greyFill

	# # R10C3
	## Oct 3 requirement removed - should not be in future
	# if (project_end_date[i].value != None) and (project_end_date[i].value > datetime.now() and (project_status[i].value == "Complete" or project_status[i].value == "Closed" or (project_status[i].value.find("Complete") != -1 and project_status[i].value.find("-") != -1))):
	# 	project_status[i].fill = greyFill
	# 	project_end_date[i].fill = redFill


	#R11C1
	if (last_updated[i].value == None and project_status[i].value != "Booked"):
		last_updated[i].fill = redFill
		project_status[i].fill = greyFill

	# R11C2
	if (last_updated[i].value != None) and (last_updated[i].value < datetime.now() - timedelta(days=14)):
		last_updated[i].fill = redFill

	# R11C3
	if (last_updated[i].value != None) and (last_updated[i].value > datetime.now()):
		last_updated[i].fill = redFill

	# R12
	if(rag_status[i].value == None):
		rag_status[i].fill = redFill


	#R13
	if (rag_status[i].value == "Amber" or rag_status[i].value == "Red") and rag_reason[i].value == None:
		rag_reason[i].fill = redFill

	#R14
	if (margin_variance[i].value > -0.05):
		margin_variance[i].fill = redFill

	#R15
	if (delivery_days_variance[i].value > -0.1):
		delivery_days_variance[i].fill = redFill

	#R16
	if (reason_overrun[i].value == None and project_status[i].value != "Booked"):
		reason_overrun[i].fill = redFill

	#R23
	if (delivery_country[i].value == None):
		delivery_country[i].fill = redFill

	#R24
	if (project_type[i].value == None):
		project_type[i].fill = redFill

	# # R25
	# if (to_expiry_date[i].value == None):
	# 	to_expiry_date[i].fill = redFill

	# R26
	if (project_category[i].value == None):
		project_category[i].fill = redFill

	# R27C1
	if (project_revenue[i].value == None):
		project_revenue[i].fill = redFill

	# R27 C2
	if (project_revenue[i].value != None and revenue_source[i].value != None and (revenue_source[i].value.find("PO") != -1 )):
		project_revenue[i].fill = redFill
		# revenue_source[i].fill = redFill

	# R27C3
	if (project_revenue[i].value != 0 and (revenue_source[i].value == None or revenue_source[i].value == "CSAT" or revenue_source[i].value == "FOC" or revenue_source[i].value == "Reclass")):
		project_revenue[i].fill = redFill


	# R28
	if (amount_invoiced[i].value != None and project_revenue[i].value != None) and (amount_invoiced[i].value > project_revenue[i].value):
		amount_invoiced[i].fill = redFill


	# R29
	if (project_reclass[i].value == None and (revenue_source[i].value =="Reclass" or revenue_source[i].value == "PO/Reclass")):
		project_reclass[i].fill = redFill
		revenue_source[i].fill = greyFill


	# R30
	if (revenue_source[i].value == None):
		revenue_source[i].fill = redFill

	# R31
	if (invoiced_method[i].value == None):
		invoiced_method[i].fill = redFill

	# R33
	## New requirement Oct 3
	if (project_stage[i].value == "Awarded") and booked_date[i].value == None:
		booked_date[i].fill = redFill
		project_stage[i].fill = greyFill

		
	# R37
	if (project_revenue[i].value == None):
		project_revenue[i].fill = redFill

	# R38
	if (project_cost_budget[i].value == None):
		project_cost_budget[i].fill = redFill

	# R48
	if (project_margin_budget[i].value == None):
		project_margin_budget[i].fill = redFill



	# R77
	if (delivery_days_budget[i].value == None):
		delivery_days_budget[i].fill = redFill

	# R78
	if (delivery_days_forecast[i].value == None):
		delivery_days_forecast[i].fill = redFill

	# R110
	if (days_budget[i].value == None):
		days_budget[i].fill = redFill

	# R111
	if (days_forecast[i].value == None):
		days_forecast[i].fill = redFill


	# R8 case 1
	if project_per_complete[i].value > 1:
		project_per_complete[i].fill = redFill
	else: #R8C3
		if (project_status[i].value == "Complete" or 
		   (project_status[i].value.find("Complete") != -1 and project_status[i].value.find("-") != -1) or (project_status[i].value == "Closed")):
			project_status[i].fill = redFill
			project_per_complete[i].fill = redFill
			
	#Req R7 case 1 & R8C2:
	if project_per_complete[i].value != None and ((int(project_per_complete[i].value) >= 1) and (project_status[i].value != "Complete" 
		or project_status[i].value !="Closed")):
		project_status[i].fill = redFill
		project_per_complete[i].fill = redFill
		

	#Req R7 case 2:
	if (project_status[i].value.find("In Progress") !=-1 or project_status[i].value.find("On Hold") !=-1 or project_status[i].value.find("Complete") !=1) and (project_status[i].value.find("-") !=-1):
		if project_stage[i].value != "Opportunity":
			project_stage[i].fill = greyFill
			project_status[i].fill = redFill
		 #No more rules for WAR

	#Other way round:
	if project_stage[i].value == "Opportunity":
		if (project_status[i].value.find("In Progress") !=-1 or project_status[i].value.find("On Hold") !=-1 or project_status[i].value.find("Complete") !=1) and (project_status[i].value.find("-") ==-1):
			project_stage[i].fill = greyFill
			project_status[i].fill = redFill	


	# R7 case 3:
	if (project_status[i].value == "In Progress" or project_status[i].value == "Booked" 
		or project_status[i].value == "On Hold" or project_status[i].value == "Complete") and project_stage[i].value != "Awarded":
		project_stage[i].fill = greyFill
		project_status[i].fill = redFill

	#other way round: (Oct 3)
	if (project_stage[i].value == "Awarded" and (project_status[i].value != "In Progress" or project_status[i].value != "Booked" or project_status[i].value != "On Hold" or project_status[i].value != "Complete")):
		project_stage[i].fill = greyFill
		project_status[i].fill = redFill

	# Oct 3 2017
	## resolving not coded comment (should have already been covered under #Req R7 case 1 & R8C2:)
	if project_stage[i].value == "Complete" and project_status[i].value != "Closed":
		project_stage[i].fill = greyFill
		project_status[i].fill = redFill

	#New requirement Oct 3 2017
	if project_status[i].value == "Booked":
		if project_start_date[i].value != None and project_start_date[i].value < datetime.now() :
			projetct_status[i].fill = greyFill
			project_start_date.fill = redFill

		if project_end_date[i].value != None and project_end_date[i].value < datetime.now():
			projetct_status[i].fill = greyFill
			project_end_date.fill = redFill


print "All done! Creating file -> updated.xlsx"
wb.save('updated.xlsx')
print "Bye!"